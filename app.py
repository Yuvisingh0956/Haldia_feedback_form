from flask import Flask, render_template, request, redirect, url_for, session
import openpyxl
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Needed for session management

# Path to the Excel file
EXCEL_FILE = 'student_data.xlsx'

# Initialize the Excel file if it doesn't exist
def init_excel_file():
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'StudentData'
        # Add header row
        sheet.append(['Name', 'Roll Number', 'Semester', 
                      '1. Does the curriculum well designed?', 
                      '2. Was the course conceptually difficult to understand?', 
                      '3. Does the curriculum has enough content for a student to acquire sufficient knowledge to secure a subject related job?', 
                      '4. Did the curriculum promote learning outcome?', 
                      '5. Did the course curriculum intellectually stimulate ?', 
                      '6. Does the curriculum design has focus on employability?', 
                      '7. Do you the think that the syllabus is adequate for GATE ?', 
                      '8. Did the syllabus provides foundation for pursuing Higher Studies/Research ?', 
                      '9. Did the subject/courses help in developing your personality?', 
                      '10.Does the syllabus has enough innovativeness and opportunities for creative thinking?',

                      '1.The faculty explained the objective of the course. Its relevance in regard to Industrial application, current development and research opportunities.', 
                      '2. The prerequisites, pertinence of the course with others and programme as a whole and the organization of the subject matter are explained.', 
                      "3. The teacher explained CO statements and its correlations with the PO's and PSO's", 
                      '4. The teacher is enthusiastic and created interest in the subject', 
                      '5. The teacher delivered the lecture lucidly', 
                      '6.The teacher emphasized on numerical problem solving / mathematical formulation etc, example and data analysis.', 
                      '7. Teacher used modern and smart teaching aids, whenever relevant.', 
                      '8. Test, Assignment and quizzes were adequate.', 
                      '9. The teacher provides opportunities for participatory learning.', 
                      '10. Your level of satisfaction with the all round contribution of the teacher',
                      
                      '1. Do you find the curriculum is well designed?', 
                      '2. Percentage of use of ICT based teaching?', 
                      '3. Does the courses conceptually difficult to understand?', 
                      '4. Does the curriculum have focus on employability?', 
                      '5. Do you think that the syllabus is adequate for GATE?', 
                      '6. Do the notices are displayed /communicated in right time?', 
                      '7. Do you think that the tutorial classes are adequate?', 
                      '8. Do you think that the tutorial classes are adequate?', 
                      '9. Are you satisfied with the examination related procedures and timely publications of results?', 
                      '10. Are you satisfied with the library (central/departmental) facilities ?',
                      '11. Are you satisfied with the official work related to students centric documentation ?', 
                      '12. Do you think that there is adequate provision for pursuing co-curricular and extra-curricular activities?', 
                      '13. Are you satisfied with the adequacy to games,sports,gym facilities ?', 
                      '14. Are you satisfied with the Industrial training/internships/placement related preparatory measures ?', 
                      '15. Are you satisfied with the training and placement activities?', 
                      '16. Are you satisfied with the campus life and facilities?', 
                      '17. Are you satisfied with the adequacy to games,sports,gym facilities ?'])
        workbook.save(EXCEL_FILE)

# Append student details to Excel file
def save_student_details(name, roll_number, semester):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['StudentData']
    
    # Check if Roll Number already exists to prevent duplicates
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == roll_number:
            raise ValueError(f"Roll Number {roll_number} already exists.")
    
    # Append the student details with empty feedback columns
    sheet.append([name, roll_number, semester] + [''] * 10)
    workbook.save(EXCEL_FILE)

# Update student feedback in Excel file
def update_feedback(roll_number, feedback):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['StudentData']
    updated = False

    # Iterate through the rows to find the matching Roll Number
    for row in sheet.iter_rows(min_row=2):
        cell_roll_number = row[1].value
        if cell_roll_number == roll_number:
            # Update Q1 to Q10 columns (columns D to M, which are indices 3 to 12)
            for i, answer in enumerate(feedback, start=4):
                row[i-1].value = answer  # openpyxl is 1-indexed
            updated = True
            break

    if not updated:
        raise ValueError(f"Roll Number {roll_number} not found.")
    
    workbook.save(EXCEL_FILE)

# Update faculty feedback in Excel file
def faculty_feedback(roll_number, feedback):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['StudentData']
    updated = False

    # Iterate through the rows to find the matching Roll Number
    for row in sheet.iter_rows(min_row=2):
        cell_roll_number = row[1].value
        if cell_roll_number == roll_number:
            # Update Q1 to Q10 columns (columns D to M, which are indices 3 to 12)
            for i, answer in enumerate(feedback, start=14):
                row[i-1].value = answer  # openpyxl is 1-indexed
            updated = True
            break

    if not updated:
        raise ValueError(f"Roll Number {roll_number} not found.")
    
    workbook.save(EXCEL_FILE)

# Update student satisfaction feedback in Excel file
def student_satisfaction(roll_number, feedback):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['StudentData']
    updated = False

    # Iterate through the rows to find the matching Roll Number
    for row in sheet.iter_rows(min_row=2):
        cell_roll_number = row[1].value
        if cell_roll_number == roll_number:
            # Update Q1 to Q10 columns (columns D to M, which are indices 3 to 12)
            for i, answer in enumerate(feedback, start=24):
                row[i-1].value = answer  # openpyxl is 1-indexed
            updated = True
            break

    if not updated:
        raise ValueError(f"Roll Number {roll_number} not found.")
    
    workbook.save(EXCEL_FILE)

@app.route('/')
def index():
    # Render the student details form first
    return render_template('student_details.html')

@app.route('/submit_details', methods=['POST'])
def submit_details():
    try:
        # Capture student details
        name = request.form['name']
        roll_number = request.form['roll_number']
        semester = request.form['semester']
        
        # Basic validation
        if not name or not roll_number or not semester:
            return "All fields are required.", 400
        
        # Save the student details to Excel
        save_student_details(name, roll_number, semester)
        
        # Store details in session for later use
        session['name'] = name
        session['roll_number'] = roll_number
        session['semester'] = semester

        # Redirect to feedback form
        return redirect(url_for('feedback_form'))
    except ValueError as ve:
        return str(ve), 400
    except Exception as e:
        return "An error occurred while saving details.", 500

@app.route('/feedback_form')
def feedback_form():
    # Ensure student details are in session
    if 'name' not in session or 'roll_number' not in session or 'semester' not in session:
        return redirect(url_for('index'))
    
    # Render the feedback form with student details
    return render_template('feedback_form.html', 
                           name=session['name'], 
                           roll_number=session['roll_number'], 
                           semester=session['semester'])

@app.route('/submit_feedback', methods=['POST'])
def submit_feedback():
    try:
        # Capture feedback responses
        feedback = []
        for i in range(1, 11):
            answer = request.form.get(f'q{i}', '')
            if not answer:
                return f"Question Q{i} is required.", 400
            feedback.append(answer)
        
        # Retrieve student details from session
        roll_number = session.get('roll_number')
        if not roll_number:
            return "Session expired or invalid.", 400
        
        # Update feedback in the Excel file
        update_feedback(roll_number, feedback)

        return render_template('faculty_feedback.html', 
                           name=session['name'], 
                           roll_number=session['roll_number'], 
                           semester=session['semester'])
    except ValueError as ve:
        return str(ve), 400
    except Exception as e:
        return "An error occurred while saving feedback.", 500

@app.route('/faculty_feedback', methods=['POST'])
def submit_faculty_feedback():
    # Ensure student details are in session
    if 'name' not in session or 'roll_number' not in session or 'semester' not in session:
        return redirect(url_for('index'))
    else:
        try:
            # Capture faculty feedback responses
            feedback = []
            for i in range(1, 11):
                answer = request.form.get(f'q{i}', '')
                if not answer:
                    return f"Question Q{i} is required.", 400
                feedback.append(answer)
            
            # Retrieve student details from session
            roll_number = session.get('roll_number')
            if not roll_number:
                return "Session expired or invalid.", 400
            
            # Update faculty feedback in the Excel file
            faculty_feedback(roll_number, feedback)

            return render_template('student_satisfaction.html', 
                           name=session['name'], 
                           roll_number=session['roll_number'], 
                           semester=session['semester'])
            
            
        except ValueError as ve:
            return str(ve), 400
        except Exception as e:
            return "An error occurred while saving feedback.", 500

@app.route('/student_satisfaction', methods=['POST'])
def submit_student_satisfaction():
    # Ensure student details are in session
    if 'name' not in session or 'roll_number' not in session or 'semester' not in session:
        return redirect(url_for('index'))
    else:
        try:
            # Capture faculty feedback responses
            feedback = []
            for i in range(1, 18):
                answer = request.form.get(f'q{i}', '')
                if not answer:
                    return f"Question Q{i} is required.", 400
                feedback.append(answer)
            
            # Retrieve student details from session
            roll_number = session.get('roll_number')
            if not roll_number:
                return "Session expired or invalid.", 400
            
            # Update faculty feedback in the Excel file
            student_satisfaction(roll_number, feedback)
            
            # Clear the session
            session.pop('name', None)
            session.pop('roll_number', None)
            session.pop('semester', None)

            return redirect(url_for('thank_you'))
        except ValueError as ve:
            return str(ve), 400
        except Exception as e:
            return "An error occurred while saving feedback.", 500

@app.route('/thank_you')
def thank_you():
    return render_template('thanku.html')


if __name__ == '__main__':
    init_excel_file()  # Initialize Excel if it doesn't exist
    app.run(debug=True)
